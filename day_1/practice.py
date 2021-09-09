import openpyxl
filename = "C:\\Users\Bill\VetsInTech\itp_week_3\day_1\lecture.xlsx"
wb = openpyxl.load_workbook(filename)
my_sheet = wb['Sheet1']

# for i in range(1, 8):
#     print(i, my_sheet.cell(row = i, column = 1).value)
#     print(i, my_sheet.cell(row = i, column = 2).value)
#     print(i, my_sheet.cell(row = i, column = 3).value)




row_count = my_sheet.max_row
column_count = my_sheet.max_column

row_count += 1
column_count += 1
print(row_count)
print(column_count)

for i in range(1, row_count):
    print(i, my_sheet.cell(row = i, column = 1).value)
    print(i, my_sheet.cell(row = i, column = 2).value)
    print(i, my_sheet.cell(row = i, column = 3).value)


# instructor solution

# import openpyxl

# wb = openpyxl.load_workbook("/home/dkayzee/vit/intro-python-august-2021/itp_week_3/day_1/lecture.xlsx")
# print(type(wb))

# sheet = wb['Sheet1']
# print(sheet.max_row) # 7

# for i in range(1, sheet.max_row + 1):
#     # on the date of A, C amount of B were sold.
#     date = "A" + str(i)
#     date_cell = sheet[date]

#     amount = "C" + str(i)
#     amount_cell = sheet[amount]

#     fruit = "B" + str(i)
#     fruit_cell = sheet[fruit]

#     print("On the Date of " + str(date_cell.value) + ", " + str(amount_cell.value) + " amount of " + fruit_cell.value + "!")