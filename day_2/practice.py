# Practice Questions:

# For the following questions, imagine you have a Workbook object in the variable wb, a Worksheet object in sheet, and a Cell object in cell.
import os
import openpyxl
from openpyxl.workbook.workbook import Workbook

# 1. What does the openpyxl.load_workbook() function return?
# open an existing workbook
wb = openpyxl.load_workbook('day_2/lecture.xlsx')

# 2. What does the wb.sheetnames workbook attribute contain?
sheet_name = wb.sheetnames

# 3. How would you retrieve the Worksheet object for a sheet named 'Sheet1'?
my_sheet = wb['Sheet1']
sheet_name = wb['Sheet1'].sheetname

# 4. How would you retrieve the Worksheet object for the workbook’s active sheet?
my_new_worksheet = my_new_workbook.active 
print(wb.active)

# 5. How would you retrieve the value in the cell C5?
cellC5Value = my_sheet['C5'].value

# 6. How would you set the value in the cell C5 to "Hello"?
my_sheet['C5'] = 'Hello'

# 7. How would you retrieve the cell’s row and column as integers?
cellC5 = my_sheet['C5']
cellRow = cellC5.row
cellColumn = cellC5.column

# 8. How would you save the workbook to the filename example.xlsx?
import os
my_new_workbook.save('example.xlsx')

# 9. If you needed to get the integer index for column 'M', what function would you need to call?

# 10. If you needed to get the string name for column 14, what function would you need to call?

# BONUS: What do the sheet.max_column and sheet.max_row sheet attributes hold, and what is the data type of these attributes?

# BONUS: How can you retrieve a tuple of all the Cell objects from A1 to F1?

# BONUS: How would you set the height of row 5 to 100?
