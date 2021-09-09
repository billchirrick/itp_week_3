import os
import openpyxl
from openpyxl import workbook
from openpyxl.workbook.workbook import Workbook

wb = openpyxl.load_workbook('day_2/lecture.xlsx')

# wb.sheetnames
# print(str(wb))


# wb = Workbook()
# print(str(wb))

# new_sheet = wb.create_sheet()
# print(str(new_sheet))

for i in range(3, 10):
    wb.create_sheet()

for sheet in wb:
    y = 50
    sheet.title = "MyNewSheets" str(y)
    y += 1

print(str(wb.sheetnames))

# from openpyxl import workbook
# for value in workbook:
#     if type(value) != int:
#         int(value)


import openpyxl
from openpyxl.workbook import workbook
from openpyxl.workbook.workbook import Workbook

wb = openpyxl.load_workbook('day_2/lecture.xlsx')

#print(str(wb.sheetnames))

my_sheet1 = wb['Population by Census Tract']
my_column = my_sheet1

for i in range(1, my_sheet1.max_row+1):
    if i == None:
        break
    i = str(i)
    print(i + "Im a string")


# As far as the cell object not being the same thing as the cell value, here's another way to think about it:
# document: {
#   Sheet1: {
#     cell1: {
#       value: 'Apples'
#       row: 1,
#       column: 1
#     }
#     cell2: {
#       value: 'Oranges',
#       row: 2,
#       column: 1
#     }
#     cell3: {
#       value: 'Grapes',
#       row: 3,
#       column: 1
#     }              #etc...
#   },  
#                       #or another way to see that object:
#   Some_Census_Data: {
#       cell1: {value: 'Georgia', row: 1, column: 1} #etc
#       cell2: {value: 'California'}
#   }
# }